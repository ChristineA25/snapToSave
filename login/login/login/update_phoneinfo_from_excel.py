
# -*- coding: utf-8 -*-
"""
Update phoneInfo.offsetHrsVSutc and phoneInfo.timezoneURL
by mapping Excel 'country2' (ISO-3) to phoneInfo.regionID.

This script self-bootstraps:
- Ensures pip is available
- Auto-installs 'pymysql' (and 'openpyxl' if missing) into the current interpreter/venv
- Performs robust header detection
- Safely converts offset values to float (fixes "ConvertibleToFloat" typing complaints)
"""

from __future__ import annotations

# -------------------- BOOTSTRAP: ensure pip + packages --------------------
import sys
import subprocess

def ensure_pip() -> None:
    """Ensure pip is available for the current interpreter."""
    try:
        import pip  # noqa: F401
        return
    except Exception:
        pass
    # Try to bootstrap pip
    subprocess.check_call([sys.executable, "-m", "ensurepip", "--default-pip"])
    # Upgrade pip (optional but recommended)
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"])
    except subprocess.CalledProcessError:
        # Not fatal; continue with whatever pip we have
        pass

def ensure_package(dist_name: str, import_name: str | None = None, version_spec: str | None = None) -> None:
    """
    Import a package; if missing, install it (optionally with a version spec), then import again.

    Args:
        dist_name: Distribution name on PyPI (e.g., "pymysql")
        import_name: Module name to import (defaults to dist_name)
        version_spec: e.g., "==1.1.1" or ">=1.1,<2"
    """
    name = import_name or dist_name
    try:
        __import__(name)
        return
    except ImportError:
        pass

    ensure_pip()
    target = f"{dist_name}{version_spec or ''}"
    subprocess.check_call([sys.executable, "-m", "pip", "install", target])

    # Try the import again (will raise if install failed)
    __import__(name)

# Ensure required packages (pymysql required; openpyxl suggested)
ensure_package("pymysql", "pymysql")       # <-- required by this script
ensure_package("openpyxl", "openpyxl")     # <-- needed to read Excel; remove if already guaranteed

# -------------------- ACTUAL SCRIPT BEGINS --------------------
import pymysql
from decimal import Decimal
from typing import Any, Optional
from openpyxl import load_workbook

# --------------- DB CONNECTION ---------------
# If you use different credentials/env, edit these values.
conn = pymysql.connect(
    host="centerbeam.proxy.rlwy.net",
    user="root",
    password="FNpMDDIVKerGZAFgoaJHalKfOmELHkQq",
    db="railway",
    port=46160,
)
cursor = conn.cursor()

# --------------- INPUT EXCEL ---------------
# Expected columns (case-insensitive):
# - country2
# - offset_hours_vs_utc
# - timezone_url
EXCEL_PATH = "country_timezone_offsets_iso3.xlsx"
SHEET_NAME = 0  # first worksheet

# Load workbook & sheet
wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb.worksheets[SHEET_NAME]

# --- Header row (safe, deterministic) ---
# Read exactly the first row as header; if missing, raise.
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
if header_row is None:
    raise ValueError("The worksheet is empty: no header row found.")

# Data starts on the next row after header
start_data_row = 2

# Map column names to indices (zero-based)
header_map = {
    str(name).strip().lower(): idx
    for idx, name in enumerate(header_row)
    if name is not None
}

required = ["country2", "offset_hours_vs_utc", "timezone_url"]
missing = [c for c in required if c not in header_map]
if missing:
    raise ValueError(f"Missing required Excel columns: {missing}")

idx_country2 = header_map["country2"]
idx_offset = header_map["offset_hours_vs_utc"]
idx_tzurl = header_map["timezone_url"]

# --------------- Helpers ---------------

def parse_offset(value: Any) -> Optional[float]:
    """
    Convert the Excel cell value for 'offset_hours_vs_utc' to a float.
    Returns:
        float value or None (if blank/invalid) so DB can store NULL.
    Handles:
        - int/float/Decimal → float
        - str → float if numeric, else None if blank/invalid
        - bool → float(value) (rare; treat True=1.0/False=0.0)
    Everything else returns None to avoid unsafe coercion.
    """
    if value is None:
        return None

    # Common numeric types
    if isinstance(value, (int, float, Decimal)):
        # Note: bool is a subclass of int; handled below if needed
        return float(value)

    # Booleans (unlikely for timezone offsets, but safe)
    if isinstance(value, bool):
        return float(value)

    # Strings: strip and attempt parse
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None

    # For all other types (dates, formulas, rich text, etc.) → skip/NULL
    return None

def normalize_str(value: Any) -> Optional[str]:
    """Return a trimmed string or None if blank/None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None

# --------------- UPDATE LOOP ---------------
updated = 0
skipped = 0
errors = 0

update_sql = """
UPDATE phoneInfo
SET offsetHrsVSutc = %s,
    timezoneURL = %s
WHERE regionID = %s;
"""

# Iterate data rows
for row in ws.iter_rows(min_row=start_data_row, values_only=True):
    if row is None:
        continue

    country2 = row[idx_country2] if idx_country2 < len(row) else None
    offset_raw = row[idx_offset] if idx_offset < len(row) else None
    tzurl_raw = row[idx_tzurl] if idx_tzurl < len(row) else None

    # Key check
    region_id = normalize_str(country2)
    if not region_id:
        skipped += 1
        continue
    region_id = region_id.lower()

    # Convert fields
    offset_param = parse_offset(offset_raw)
    tzurl_param = normalize_str(tzurl_raw)

    try:
        cursor.execute(update_sql, (offset_param, tzurl_param, region_id))
        if cursor.rowcount > 0:
            updated += cursor.rowcount
        else:
            # No matching regionID
            skipped += 1
    except Exception as e:
        errors += 1
        print(f"[ERROR] regionID={region_id}: {e}")

# --------------- COMMIT & CLOSE ---------------
conn.commit()
cursor.close()
conn.close()

print(f"Done. Updated rows: {updated}, Skipped (no match/blank): {skipped}, Errors: {errors}")
